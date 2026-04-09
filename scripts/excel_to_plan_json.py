#!/usr/bin/env python3
"""
Excel SOP → v2 JSON 转换脚本

将固定格式的 SOP Excel 文件转换为系统可导入的 JSON 格式。
基于首钢医护人员社群运营SOP-培训表.xlsx 的格式设计。

用法:
  python scripts/excel_to_plan_json.py input.xlsx -o output.json
  python scripts/excel_to_plan_json.py input.xlsx --validate ref.json
  python scripts/excel_to_plan_json.py input.xlsx -o output.json --scope "第二阶段：XXX"
"""

from __future__ import annotations

import argparse
import io
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

# Windows 终端 UTF-8 输出
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ── 常量 ──────────────────────────────────────────────────────────────────

# 0-indexed 列号 → JSON 字段 + stage 元信息
COLUMN_MAP = [
    {"col": 3, "field": "morning_text",                "stage_key": "morning_greeting_points", "stage_name": "早餐提醒+早安+积分发布"},
    {"col": 4, "field": "pre_lunch_knowledge_text",    "stage_key": "pre_lunch_knowledge",     "stage_name": "午餐前：知识点预告-讲科学+给方法"},
    {"col": 5, "field": "lunch_reminder_text",          "stage_key": "lunch_reminder",          "stage_name": "午餐提醒"},
    {"col": 6, "field": "afternoon_meal_review_text",   "stage_key": "afternoon_meal_review",   "stage_name": "下午餐评+互动+答疑"},
    {"col": 7, "field": "dinner_reminder_text",         "stage_key": "dinner_reminder",         "stage_name": "晚餐提醒+答疑"},
    {"col": 8, "field": "dinner_meal_review_text",      "stage_key": "dinner_meal_review",      "stage_name": "晚餐餐评+互动+答疑"},
    {"col": 9, "field": "night_summary_text",           "stage_key": "night_summary_preview",   "stage_name": "晚安+总结+明天预告"},
]

DAY_PATTERN = re.compile(r"^第([一二三四五六七八九十百零两\d]+)天$")
WEEK_PATTERN = re.compile(r"^第([一二三四五六七八九十百零两\d]+)周$")

# 默认模板值（从 Day 1 提取，用于填充缺失列）
DEFAULT_LUNCH_REMINDER = "午餐时间，我来提醒群内各位伙伴，拍照晒餐哦~"
DEFAULT_AFTERNOON_REVIEW = "餐评模板见Agent后台"
DEFAULT_DINNER_REMINDER = "晚餐时间，我来提醒群内各位老师，拍照晒餐哦~"
DEFAULT_DINNER_REVIEW = "餐评模板见Agent后台"


# ── 工具函数 ──────────────────────────────────────────────────────────────

def normalize(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def chinese_to_int(value: str) -> int:
    raw = value.strip()
    if raw.isdigit():
        return int(raw)
    mapping = {"零": 0, "一": 1, "二": 2, "两": 2, "三": 3, "四": 4,
               "五": 5, "六": 6, "七": 7, "八": 8, "九": 9}
    if raw == "十":
        return 10
    if "十" in raw:
        left, right = raw.split("十", 1)
        tens = mapping.get(left, 1 if left == "" else 0)
        ones = mapping.get(right, 0) if right else 0
        return tens * 10 + ones
    total = 0
    for char in raw:
        total = total * 10 + mapping[char]
    return total


# ── 表头检测 ──────────────────────────────────────────────────────────────

def find_header_row(rows: list[list[Any]]) -> int:
    """找到表头行（col0='每周', col3 含 '早餐提醒'）"""
    for idx, row in enumerate(rows):
        vals = [normalize(c) for c in row]
        if len(vals) >= 4 and vals[0] == "每周" and "早餐提醒" in vals[3]:
            return idx
    raise ValueError("未识别到标准表头行，请确认 Excel 格式是否正确")


# ── 阶段模板提取 ──────────────────────────────────────────────────────────

def extract_stage_templates(rows: list[list[Any]], header_idx: int) -> list[dict]:
    """从表头行下方的流程说明行和时间行提取 stage_templates"""
    workflow_row = rows[header_idx + 1] if header_idx + 1 < len(rows) else []
    time_row = rows[header_idx + 2] if header_idx + 2 < len(rows) else []

    templates = []
    for order, cmap in enumerate(COLUMN_MAP, start=1):
        col = cmap["col"]
        scheduled_time = normalize(time_row[col]) if col < len(time_row) else ""
        # 解析时间行可能包含合并的多个时间段，如 "8：00" 或 "11：3012：00-17：0017：3020：00"
        if scheduled_time:
            scheduled_time = scheduled_time.replace("：", ":")
        fixed_action = normalize(workflow_row[col]) if col < len(workflow_row) else ""
        templates.append({
            "stage_order": order,
            "stage_key": cmap["stage_key"],
            "stage_name": cmap["stage_name"],
            "scheduled_time": scheduled_time,
            "time_confidence": "high" if order <= 2 else "medium",
            "owner": "",
            "fixed_action": fixed_action,
            "source_basis": f"主表第{header_idx + 4}-{header_idx + 5}行",
        })
    return templates


# ── 右侧附加信息提取 ──────────────────────────────────────────────────────

def extract_side_info(rows: list[list[Any]], header_idx: int) -> dict:
    """从 col 11-13 提取 weekly_events 和 operator_fixed_rules"""
    weekly_events = []
    operator_rules: list[str] = []

    data_start = header_idx + 3
    for row in rows[data_start:data_start + 10]:
        if len(row) <= 12:
            continue
        col11 = normalize(row[11])
        col12 = normalize(row[12])
        col13 = normalize(row[13]) if len(row) > 13 else ""

        # 每周X + 事件
        if col11.startswith("每周") and col12 and col13:
            weekly_events.append({
                "event_key": "",
                "day_rule": col11,
                "time": "未写明确切时间",
                "owner": "",
                "action": col13,
            })
            if col12 not in (col11, col13):
                weekly_events[-1]["owner"] = col12

        # 医护必做动作
        if "必做动作" in col11 or "考核标准" in col11:
            rules_text = col12 or col13
            if rules_text:
                for line in rules_text.split("\n"):
                    line = line.strip()
                    if line and re.match(r"^\d+\.", line):
                        operator_rules.append(re.sub(r"^\d+\.\s*", "", line))

    return {"weekly_events": weekly_events, "operator_fixed_rules": operator_rules}


# ── 核心转换逻辑 ──────────────────────────────────────────────────────────

def convert_excel_to_json(excel_path: str, scope_override: str = "") -> dict:
    """将 Excel 文件转换为 v2 JSON 格式"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]

    if not rows:
        raise ValueError("Excel 文件为空")

    # 1. 提取 extraction_scope
    title_cell = normalize(rows[0][0]) if rows[0] else ""
    extraction_scope = scope_override or title_cell

    # 2. 找到表头
    header_idx = find_header_row(rows)

    # 3. stage_templates
    stage_templates = extract_stage_templates(rows, header_idx)

    # 4. 右侧附加信息
    side = extract_side_info(rows, header_idx)

    # 5. 遍历数据行，提取每天内容
    week_label = ""
    current_day: dict[str, Any] | None = None
    days: list[dict[str, Any]] = []

    for row in rows[header_idx + 3:]:
        col0 = normalize(row[0]) if len(row) > 0 else ""
        col1 = normalize(row[1]) if len(row) > 1 else ""
        col2 = normalize(row[2]) if len(row) > 2 else ""

        # 周标记
        week_match = WEEK_PATTERN.match(col0)
        if week_match:
            week_label = col0
            # 注意：不要 continue，同一行可能还有 "第X天"
            if not col1:
                continue

        # 天标记 — 开始新的一天
        day_match = DAY_PATTERN.match(col1)
        if day_match:
            if current_day:
                days.append(current_day)
            day_num = chinese_to_int(day_match.group(1))
            current_day = {
                "day": day_num,
                "week_label": week_label,
                "phase_name": "",
                "required_actions": [],
                "knowledge_topic": "",
                "morning_text": "",
                "pre_lunch_knowledge_text": "",
                "knowledge_extra_text": None,
                "knowledge_evidence_text": None,
                "lunch_reminder_text": "",
                "afternoon_meal_review_text": "",
                "dinner_reminder_text": "",
                "dinner_meal_review_text": "",
                "night_summary_text": "",
            }
            # 填充主行内容
            for cmap in COLUMN_MAP:
                col = cmap["col"]
                text = normalize(row[col]) if col < len(row) else ""
                current_day[cmap["field"]] = text
            continue

        # 续行（col1 为空，但有 col2 标签或内容）
        if current_day and col2:
            is_knowledge_label = any(kw in col2 for kw in ["科普知识点", "科学机制"])
            is_media_label = any(kw in col2 for kw in ["图片", "视频", "资料"])

            if is_knowledge_label:
                # col 4 → knowledge_extra_text, col 5 → knowledge_evidence_text
                extra = normalize(row[4]) if len(row) > 4 else ""
                evidence = normalize(row[5]) if len(row) > 5 else ""
                if extra:
                    current_day["knowledge_extra_text"] = (
                        (current_day["knowledge_extra_text"] or "") + "\n\n" + extra
                    ).strip()
                if evidence:
                    current_day["knowledge_evidence_text"] = (
                        (current_day["knowledge_evidence_text"] or "") + "\n\n" + evidence
                    ).strip()
            elif not is_media_label:
                # 其他标签行：内容追加到对应列主字段
                for cmap in COLUMN_MAP:
                    col = cmap["col"]
                    text = normalize(row[col]) if col < len(row) else ""
                    if text:
                        block = f"{col2}\n{text}" if col2 else text
                        current_day[cmap["field"]] = (
                            current_day[cmap["field"]] + "\n\n" + block
                        ).strip()

        # col1/col2 都为空但列 3-9 有内容的续行
        elif current_day and not col1 and not col2:
            for cmap in COLUMN_MAP:
                col = cmap["col"]
                text = normalize(row[col]) if col < len(row) else ""
                if text:
                    current_day[cmap["field"]] = (
                        current_day[cmap["field"]] + "\n\n" + text
                    ).strip()

    if current_day:
        days.append(current_day)

    # 填充缺失的午餐/晚餐提醒模板（从 Day 1 传播到所有天）
    for day in days:
        if not day["lunch_reminder_text"]:
            day["lunch_reminder_text"] = DEFAULT_LUNCH_REMINDER
        if not day["afternoon_meal_review_text"]:
            day["afternoon_meal_review_text"] = DEFAULT_AFTERNOON_REVIEW
        if not day["dinner_reminder_text"]:
            day["dinner_reminder_text"] = DEFAULT_DINNER_REMINDER
        if not day["dinner_meal_review_text"]:
            day["dinner_meal_review_text"] = DEFAULT_DINNER_REVIEW

    # 清理 null 字段（为 None 的改为 null，为空字符串保持）
    for day in days:
        for key in ("knowledge_extra_text", "knowledge_evidence_text"):
            if day[key] is not None and not day[key].strip():
                day[key] = None

    # 6. 组装输出
    result = {
        "source_file": Path(excel_path).name,
        "extraction_scope": extraction_scope,
        "important_notes": [
            f"由脚本自动从 Excel 提取，共 {len(days)} 天。",
        ],
        "operator_fixed_rules": side["operator_fixed_rules"],
        "stage_templates": stage_templates,
        "weekly_events": side["weekly_events"],
        "phase_rules": [],
        "days": sorted(days, key=lambda d: d["day"]),
    }

    return result


# ── 验证模式 ──────────────────────────────────────────────────────────────

def validate_against_ref(generated: dict, ref_path: str) -> None:
    """对比生成的 JSON 与参考 JSON，输出差异报告"""
    with open(ref_path, "r", encoding="utf-8") as f:
        ref = json.load(f)

    print("=" * 70)
    print("验证报告：生成的 JSON vs 参考文件")
    print("=" * 70)

    # 对比顶层字段
    for key in ("extraction_scope", "source_file"):
        gen_val = generated.get(key, "")
        ref_val = ref.get(key, "")
        status = "✓" if gen_val == ref_val else "✗"
        if status == "✗":
            print(f"\n[{status}] {key}")
            print(f"  生成: {gen_val[:80]}")
            print(f"  参考: {ref_val[:80]}")

    # 对比天数
    gen_days = {d["day"]: d for d in generated["days"]}
    ref_days = {d["day"]: d for d in ref["days"]}

    total_fields = 0
    match_fields = 0
    mismatch_details: list[str] = []

    text_fields = [
        "morning_text", "pre_lunch_knowledge_text", "knowledge_extra_text",
        "knowledge_evidence_text", "lunch_reminder_text",
        "afternoon_meal_review_text", "dinner_reminder_text",
        "dinner_meal_review_text", "night_summary_text",
    ]

    for day_num in sorted(ref_days.keys()):
        if day_num not in gen_days:
            print(f"\n[✗] 第 {day_num} 天：在生成结果中缺失")
            continue

        g = gen_days[day_num]
        r = ref_days[day_num]

        for field in text_fields:
            g_val = (g.get(field) or "").strip()
            r_val = (r.get(field) or "").strip()

            # 处理 None
            if g_val == "" and r_val == "" and g.get(field) is None and r.get(field) is None:
                total_fields += 1
                match_fields += 1
                continue

            total_fields += 1
            if g_val == r_val:
                match_fields += 1
            else:
                mismatch_details.append(
                    f"第{day_num}天 {field}: "
                    f"生成=[{g_val[:60]}...] 参考=[{r_val[:60]}...]"
                )

    # 输出差异摘要
    print(f"\n{'─' * 70}")
    print(f"总计 {total_fields} 个文本字段，匹配 {match_fields}，差异 {total_fields - match_fields}")
    accuracy = match_fields / total_fields * 100 if total_fields else 0
    print(f"匹配率: {accuracy:.1f}%")

    if mismatch_details:
        print(f"\n差异明细（共 {len(mismatch_details)} 项）：")
        for detail in mismatch_details[:30]:
            print(f"  ✗ {detail}")
        if len(mismatch_details) > 30:
            print(f"  ... 还有 {len(mismatch_details) - 30} 项差异")

    # 对比 stage_templates
    gen_stages = {s["stage_key"]: s for s in generated.get("stage_templates", [])}
    ref_stages = {s["stage_key"]: s for s in ref.get("stage_templates", [])}
    if gen_stages.keys() == ref_stages.keys():
        print(f"\n✓ stage_templates: {len(gen_stages)} 个阶段全部匹配")
    else:
        missing = ref_stages.keys() - gen_stages.keys()
        extra = gen_stages.keys() - ref_stages.keys()
        if missing:
            print(f"\n✗ stage_templates 缺少: {missing}")
        if extra:
            print(f"\n✗ stage_templates 多出: {extra}")

    print("=" * 70)


# ── CLI 入口 ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Excel SOP → v2 JSON 转换")
    parser.add_argument("excel", help="输入的 Excel 文件路径")
    parser.add_argument("-o", "--output", help="输出 JSON 文件路径")
    parser.add_argument("--validate", help="参考 JSON 文件路径（用于对比验证）")
    parser.add_argument("--scope", help="手动指定 extraction_scope（覆盖自动检测）")
    args = parser.parse_args()

    if not Path(args.excel).exists():
        print(f"错误：文件不存在 {args.excel}", file=sys.stderr)
        sys.exit(1)

    result = convert_excel_to_json(args.excel, scope_override=args.scope or "")

    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"已输出到 {args.output}")

    if args.validate:
        validate_against_ref(result, args.validate)

    if not args.output and not args.validate:
        json.dump(result, sys.stdout, ensure_ascii=False, indent=2)
        print()


if __name__ == "__main__":
    main()
