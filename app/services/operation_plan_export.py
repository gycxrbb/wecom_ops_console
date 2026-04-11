"""运营编排导出服务：Plan → JSON / Excel（100% 复刻参考模板）"""
from __future__ import annotations

import json
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.services.operation_plan_import import (
    JSON_STAGE_RULE_MAP,
    NODE_COLUMN_RULES,
)

# node_type → 友好标题
_NODE_TYPE_TITLES: dict[str, str] = {}
for _rule in NODE_COLUMN_RULES:
    _NODE_TYPE_TITLES[_rule["node_type"]] = _rule["title"]
for _mapping in JSON_STAGE_RULE_MAP.values():
    _NODE_TYPE_TITLES.setdefault(_mapping["node_type"], _mapping["title"])


def _get_node_title(node_type: str) -> str:
    return _NODE_TYPE_TITLES.get(node_type, node_type)


# ── 7 个标准动作列定义 ──────────────────────────────────────────────────
# 每个标准列用一组关键词匹配 title，node 匹配到哪个列就放到哪一列
STANDARD_COLUMNS = [
    {
        "key": "morning",
        "header": "早餐提醒+早安+积分发布",
        "keywords": ["早餐", "早安", "积分"],
        "node_types": {"morning_breakfast", "score_publish"},
    },
    {
        "key": "before_lunch",
        "header": "午餐前：知识点预告-讲科学+给方法",
        "keywords": ["午餐前", "知识点预告"],
        "node_types": {"before_lunch_content"},
    },
    {
        "key": "lunch",
        "header": "午餐提醒",
        "keywords": ["午餐提醒", "午餐时间"],
        "node_types": {"lunch_reminder"},
        "exclude_keywords": ["科普", "知识点推送", "视频"],
    },
    {
        "key": "afternoon",
        "header": "下午餐评+互动+答疑",
        "keywords": ["下午", "餐评"],
        "node_types": {"afternoon_review"},
    },
    {
        "key": "dinner",
        "header": "晚餐提醒+答疑",
        "keywords": ["晚餐提醒"],
        "node_types": {"dinner_reminder"},
        "exclude_keywords": ["餐评"],
    },
    {
        "key": "evening",
        "header": "晚餐餐评+互动+答疑",
        "keywords": ["晚餐餐评", "晚上"],
        "node_types": {"evening_review"},
    },
    {
        "key": "night",
        "header": "晚安+总结+明天预告",
        "keywords": ["晚安", "总结", "预告"],
        "node_types": {"night_summary"},
    },
]

# msg_type → C列子行标签
_MSG_TYPE_LABELS: dict[str, str] = {
    "text": "实用知识点",
    "markdown": "实用知识点",
    "image": "图片/视频",
    "file": "图片/视频",
    "news": "图文链接",
    "template_card": "模板卡片",
}


def _match_column(node: dict) -> str | None:
    """将一个节点匹配到标准列 key，未匹配返回 None"""
    nt = node.get("node_type", "")
    title = node.get("title", "")

    for col_def in STANDARD_COLUMNS:
        # 优先按 node_type 匹配
        if nt in col_def["node_types"]:
            # 排除关键词检查
            excl = col_def.get("exclude_keywords", [])
            if any(kw in title for kw in excl):
                continue
            return col_def["key"]

    # 兜底：按 title 关键词匹配
    for col_def in STANDARD_COLUMNS:
        if any(kw in title for kw in col_def["keywords"]):
            excl = col_def.get("exclude_keywords", [])
            if any(kw in title for kw in excl):
                continue
            return col_def["key"]

    return None


def _sub_row_label(msg_type: str) -> str:
    return _MSG_TYPE_LABELS.get(msg_type, "实用知识点")


# ── 通用序列化 ────────────────────────────────────────────────────────────

def export_plan_to_dict(plan: Any) -> dict[str, Any]:
    """将 Plan ORM 对象序列化为通用 dict（可 JSON 化）"""
    days_out: list[dict[str, Any]] = []
    for day in sorted(plan.days, key=lambda d: d.day_number):
        nodes_out: list[dict[str, Any]] = []
        for node in sorted(day.nodes, key=lambda n: n.sort_order):
            content_json = node.content_json
            if isinstance(content_json, str):
                try:
                    content_json = json.loads(content_json)
                except (json.JSONDecodeError, TypeError):
                    content_json = {}
            variables_json = node.variables_json
            if isinstance(variables_json, str):
                try:
                    variables_json = json.loads(variables_json)
                except (json.JSONDecodeError, TypeError):
                    variables_json = {}

            nodes_out.append({
                "node_type": node.node_type or "custom",
                "title": node.title or "",
                "msg_type": node.msg_type or "markdown",
                "description": node.description or "",
                "content": content_json.get("content", "") if isinstance(content_json, dict) else str(content_json),
                "sort_order": node.sort_order,
                "variables_json": variables_json if isinstance(variables_json, dict) else {},
                "asset_name": content_json.get("asset_name", "") if isinstance(content_json, dict) else "",
            })

        days_out.append({
            "day": day.day_number,
            "week_label": _infer_week_label(day.day_number),
            "day_title": day.title or f"第{day.day_number}天",
            "day_focus": day.focus or "",
            "nodes": nodes_out,
        })

    seen_types: dict[str, dict[str, Any]] = {}
    for day_data in days_out:
        for node_data in day_data["nodes"]:
            nt = node_data["node_type"]
            if nt not in seen_types:
                seen_types[nt] = {
                    "stage_key": nt,
                    "stage_name": node_data["title"] or _get_node_title(nt),
                    "stage_order": node_data["sort_order"],
                    "description": node_data.get("description", ""),
                }

    stage_templates = sorted(seen_types.values(), key=lambda s: s["stage_order"])

    return {
        "plan_name": plan.name or "",
        "stage": plan.stage or "",
        "topic": plan.topic or "",
        "description": plan.description or "",
        "stage_templates": stage_templates,
        "days": days_out,
    }


def _infer_week_label(day_number: int) -> str:
    week = (day_number - 1) // 7 + 1
    week_map = {1: "第一周", 2: "第二周", 3: "第三周", 4: "第四周", 5: "第五周"}
    return week_map.get(week, f"第{week}周")


# ── JSON 导出 ─────────────────────────────────────────────────────────────

def export_plan_to_json_bytes(plan: Any) -> bytes:
    data = export_plan_to_dict(plan)
    return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")


# ── Excel 导出（复刻参考模板）─────────────────────────────────────────────

_THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
_CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)

_FILL_TITLE = PatternFill(start_color='FF495700', end_color='FF495700', fill_type='solid')
_FONT_TITLE = Font(name='微软雅黑', bold=True, size=13.5, color='FFF8F9FA')

_FILL_RULE = PatternFill(start_color='FFD9F5D6', end_color='FFD9F5D6', fill_type='solid')
_FONT_RULE = Font(name='微软雅黑', bold=True, size=10)

_FILL_COL_HEADER = PatternFill(start_color='FFECE2FE', end_color='FFECE2FE', fill_type='solid')
_FONT_COL_HEADER = Font(name='微软雅黑', bold=True, size=10)

_FILL_FLOW = PatternFill(start_color='FFFFF258', end_color='FFFFF258', fill_type='solid')
_FONT_FLOW = Font(name='微软雅黑', bold=True, size=10)

_FILL_TIME = PatternFill(start_color='FFD9F5D6', end_color='FFD9F5D6', fill_type='solid')

_FILL_EVEN = PatternFill(start_color='FFD9F5D6', end_color='FFD9F5D6', fill_type='solid')
_FILL_ODD = PatternFill(fill_type=None)


def _cell(ws, row: int, col: int, value: Any = None,
          font: Font | None = None, fill: PatternFill | None = None,
          alignment: Alignment | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    cell.border = _THIN_BORDER
    return cell


def _build_day_sub_rows(day_nodes: list[dict], col_key_map: dict[str, int]) -> list[dict]:
    """
    将一天的所有节点按标准列分组，再按 msg_type 拆成子行。

    策略：
    1. 先识别"锚点节点"——能明确匹配到标准列的文字/markdown节点
    2. 未匹配的节点（custom image/file）归入其前一个锚点所在的列
    3. 同一列内按 msg_type 分组为不同子行
    """
    # 按原始顺序处理节点，维护"当前列"
    # col_key → [node1, node2, ...]
    col_nodes: dict[str, list[dict]] = {}
    current_col_key: str | None = None

    for node in day_nodes:
        col_key = _match_column(node)
        if col_key and col_key in col_key_map:
            current_col_key = col_key
        # current_col_key 可能仍为 None（第一个节点就不匹配）
        if current_col_key and current_col_key in col_key_map:
            col_nodes.setdefault(current_col_key, []).append(node)
        else:
            # 兜底：放到第一个标准列
            first_key = next(iter(col_key_map), None)
            if first_key:
                col_nodes.setdefault(first_key, []).append(node)

    # 按 msg_type 拆子行: col_key → {sub_label: content}
    sub_rows_map: dict[str, dict[str, str]] = {}
    for col_key, nodes in col_nodes.items():
        for node in nodes:
            msg_type = node.get("msg_type", "markdown")
            label = _sub_row_label(msg_type)
            content = node.get("content", "")
            # 图片/文件节点优先用 asset_name（完整文件名如 xxx.jpg / xxx.mp4）
            if msg_type in ("image", "file"):
                asset_name = node.get("asset_name", "")
                if asset_name:
                    content = asset_name
                elif not content:
                    content = node.get("title", "")
            sub_rows_map.setdefault(label, {})[col_key] = content

    # 按优先级顺序输出
    result: list[dict] = []
    priority = ["实用知识点", "［科普知识点］", "【科普图片/视频资料】", "图片/视频", "图文链接", "模板卡片"]
    seen: set[str] = set()
    for label in priority:
        if label in sub_rows_map:
            result.append({"label": label, "contents": sub_rows_map[label]})
            seen.add(label)
    for label, contents in sub_rows_map.items():
        if label not in seen:
            result.append({"label": label, "contents": contents})

    if not result:
        result.append({"label": "实用知识点", "contents": {}})

    return result


def export_plan_to_excel_bytes(plan: Any) -> bytes:
    """导出为 Excel bytes，复刻参考模板格式"""
    data = export_plan_to_dict(plan)
    days = data["days"]
    if not days:
        raise ValueError("计划中没有可导出的天数")

    # ── 7 个标准列 → Excel 列索引 (D=4 起) ──
    col_key_map: dict[str, int] = {}
    for idx, col_def in enumerate(STANDARD_COLUMNS):
        col_key_map[col_def["key"]] = 4 + idx
    total_cols = 3 + len(STANDARD_COLUMNS)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "社群运营SOP"

    # ═══ Row 1: 标题行 ═══
    day_count = len(days)
    week_count = (day_count - 1) // 7 + 1
    plan_name = data["plan_name"]
    title_text = f"{plan_name}（共 {week_count} 周・{day_count} 个工作日）" if plan_name else f"运营编排（共 {week_count} 周・{day_count} 天）"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    _cell(ws, 1, 1, title_text, font=_FONT_TITLE, fill=_FILL_TITLE, alignment=_CENTER_WRAP)
    for c in range(2, total_cols + 1):
        ws.cell(row=1, column=c).fill = _FILL_TITLE
        ws.cell(row=1, column=c).border = _THIN_BORDER
    ws.row_dimensions[1].height = 36

    # ═══ Row 2: 核心规则 ═══
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    _cell(ws, 2, 1, "核心规则", font=_FONT_RULE, fill=_FILL_RULE, alignment=_CENTER_WRAP)
    for c in (2, 3):
        ws.cell(row=2, column=c).fill = _FILL_RULE
        ws.cell(row=2, column=c).border = _THIN_BORDER

    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=total_cols)
    _cell(ws, 2, 4, "每周周一～周五发当日科普（5 天 / 周）；每周五增加：本周总结 + 下周主题预告",
          font=_FONT_RULE, fill=_FILL_RULE, alignment=_CENTER_WRAP)
    for c in range(5, total_cols + 1):
        ws.cell(row=2, column=c).fill = _FILL_RULE
        ws.cell(row=2, column=c).border = _THIN_BORDER
    ws.row_dimensions[2].height = 24

    # ═══ Row 3: 列标题 ═══
    _cell(ws, 3, 1, "每周", font=_FONT_COL_HEADER, fill=_FILL_COL_HEADER, alignment=_CENTER_WRAP)
    _cell(ws, 3, 2, "时间\n（第x天）", font=_FONT_COL_HEADER, fill=_FILL_COL_HEADER, alignment=_CENTER_WRAP)
    _cell(ws, 3, 3, "节点说明", font=_FONT_COL_HEADER, fill=_FILL_COL_HEADER, alignment=_CENTER_WRAP)
    for col_def in STANDARD_COLUMNS:
        _cell(ws, 3, col_key_map[col_def["key"]], col_def["header"],
              font=_FONT_COL_HEADER, fill=_FILL_COL_HEADER, alignment=_CENTER_WRAP)
    ws.row_dimensions[3].height = 28

    # ═══ Row 4: 流程说明 ═══
    _cell(ws, 4, 1, None, fill=_FILL_FLOW)
    _cell(ws, 4, 2, "每个动作的流程说明", font=_FONT_FLOW, fill=_FILL_FLOW, alignment=_CENTER_WRAP)
    _cell(ws, 4, 3, None, fill=_FILL_FLOW)
    for col_def in STANDARD_COLUMNS:
        _cell(ws, 4, col_key_map[col_def["key"]], None, fill=_FILL_FLOW, alignment=_LEFT_WRAP)

    # ═══ Row 5: 时间行 ═══
    _cell(ws, 5, 1, None, fill=_FILL_TIME)
    _cell(ws, 5, 2, None, fill=_FILL_TIME)
    _cell(ws, 5, 3, "时间", font=_FONT_COL_HEADER, fill=_FILL_TIME, alignment=_CENTER_WRAP)
    for col_def in STANDARD_COLUMNS:
        _cell(ws, 5, col_key_map[col_def["key"]], None, fill=_FILL_TIME)

    # ═══ Row 6+: 数据区 ═══
    current_row = 6
    font_bold = Font(name='微软雅黑', bold=True, size=10)

    # 按周分组
    weeks: list[tuple[str, list[dict]]] = []
    cur_week = None
    cur_days: list[dict] = []
    for day_data in days:
        wl = day_data.get("week_label", _infer_week_label(day_data["day"]))
        if wl != cur_week:
            if cur_days:
                weeks.append((cur_week, cur_days))
            cur_week = wl
            cur_days = []
        cur_days.append(day_data)
    if cur_days:
        weeks.append((cur_week, cur_days))

    for week_label, week_days in weeks:
        week_start_row = current_row

        for day_data in week_days:
            day_num = day_data["day"]
            is_even = day_num % 2 == 0
            day_fill = _FILL_EVEN if is_even else _FILL_ODD

            # 将节点拆成子行
            sub_rows = _build_day_sub_rows(day_data["nodes"], col_key_map)
            day_start = current_row

            for sub in sub_rows:
                # C 列子行标签
                _cell(ws, current_row, 3, sub["label"], font=font_bold, fill=day_fill, alignment=_CENTER_WRAP)
                # D~J: 各标准列内容
                for col_def in STANDARD_COLUMNS:
                    key = col_def["key"]
                    content = sub["contents"].get(key, "")
                    _cell(ws, current_row, col_key_map[key], content, fill=day_fill, alignment=_LEFT_WRAP)
                # A 列占位
                _cell(ws, current_row, 1, None, fill=day_fill, alignment=_CENTER_WRAP)
                # B 列占位（后续合并写天次）
                _cell(ws, current_row, 2, None, fill=day_fill, alignment=_CENTER_WRAP)
                current_row += 1

            # B 列合并写"第x天"
            if current_row - 1 >= day_start:
                if current_row - 1 > day_start:
                    ws.merge_cells(start_row=day_start, start_column=2, end_row=current_row - 1, end_column=2)
                _cell(ws, day_start, 2, f"第{day_num}天", font=font_bold, fill=day_fill, alignment=_CENTER_WRAP)

        # A 列合并写周次
        if current_row - 1 >= week_start_row:
            if current_row - 1 > week_start_row:
                ws.merge_cells(start_row=week_start_row, start_column=1, end_row=current_row - 1, end_column=1)
            _cell(ws, week_start_row, 1, week_label, font=_FONT_COL_HEADER, fill=_FILL_COL_HEADER, alignment=_CENTER_WRAP)
            for r in range(week_start_row, current_row):
                ws.cell(row=r, column=1).border = _THIN_BORDER

    # ═══ 列宽 ═══
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    for col_def in STANDARD_COLUMNS:
        col_letter = openpyxl.utils.get_column_letter(col_key_map[col_def["key"]])
        ws.column_dimensions[col_letter].width = 40

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
