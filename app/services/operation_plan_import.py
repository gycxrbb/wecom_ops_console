from __future__ import annotations

import json
import re
from io import BytesIO
from typing import Any

import openpyxl
from fastapi import HTTPException, UploadFile

NODE_COLUMN_RULES = [
    {"column": 4, "node_type": "morning_breakfast", "title": "早安 / 早餐提醒", "msg_type": "markdown", "sort_order": 10},
    {"column": 5, "node_type": "before_lunch_content", "title": "午餐前内容发布", "msg_type": "markdown", "sort_order": 20},
    {"column": 6, "node_type": "lunch_reminder", "title": "午餐提醒发布", "msg_type": "markdown", "sort_order": 30},
    {"column": 7, "node_type": "afternoon_review", "title": "下午餐评 / 互动 / 答疑", "msg_type": "markdown", "sort_order": 40},
    {"column": 8, "node_type": "dinner_reminder", "title": "晚餐提醒发布", "msg_type": "markdown", "sort_order": 50},
    {"column": 9, "node_type": "evening_review", "title": "晚餐餐评 / 互动 / 答疑", "msg_type": "markdown", "sort_order": 60},
    {"column": 10, "node_type": "night_summary", "title": "晚安 / 总结 / 次日预告", "msg_type": "markdown", "sort_order": 70},
]

# 默认模板内容：Excel 空列时用这些内容填充
DEFAULT_NODE_CONTENT = {
    "lunch_reminder": "午餐时间，我来提醒群内各位伙伴，拍照晒餐哦~",
    "afternoon_review": "餐评模板见Agent后台",
    "dinner_reminder": "晚餐时间，我来提醒群内各位老师，拍照晒餐哦~",
    "evening_review": "餐评模板见Agent后台",
}

DAY_PATTERN = re.compile(r"^第([一二三四五六七八九十百零两\d]+)天$")
WEEK_PATTERN = re.compile(r"^第([一二三四五六七八九十百零两\d]+)周$")


def chinese_to_int(value: str) -> int:
    raw = value.strip()
    if raw.isdigit():
        return int(raw)
    mapping = {"零": 0, "一": 1, "二": 2, "两": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9}
    if raw == "十":
        return 10
    if "十" in raw:
        left, right = raw.split("十", 1)
        tens = mapping.get(left, 1 if left == "" else 0)
        ones = mapping.get(right, 0) if right else 0
        return tens * 10 + ones
    total = 0
    for char in raw:
        total = total * 10 + mapping[char]
    return total


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def parse_stage_and_topic(sheet_name: str, title_cell: str = "") -> tuple[str, str, str]:
    name = normalize_text(title_cell) or normalize_text(sheet_name)
    if name.startswith("已恢复_"):
        name = normalize_text(title_cell) or normalize_text(sheet_name)
    compact_name = re.sub(r"[-－—].*$", "", name).strip()
    if "：" in compact_name:
        stage, topic = compact_name.split("：", 1)
    elif ":" in compact_name:
        stage, topic = compact_name.split(":", 1)
    else:
        stage, topic = compact_name, compact_name
    return compact_name, stage.strip(), topic.strip()


def detect_header_row(rows: list[list[Any]]) -> int:
    for idx, row in enumerate(rows):
        values = [normalize_text(cell) for cell in row]
        if len(values) >= 4 and values[0] == "每周" and "早餐提醒" in values[3]:
            return idx
    raise HTTPException(400, "未识别到 sheet1 的标准表头，请确认 SOP 模板结构未发生变化")


def parse_sheet1_rows(rows: list[list[Any]], sheet_name: str) -> dict[str, Any]:
    header_idx = detect_header_row(rows)
    workflow_row = rows[header_idx + 1] if header_idx + 1 < len(rows) else []
    time_row = rows[header_idx + 2] if header_idx + 2 < len(rows) else []

    week_label = ""
    current_day: dict[str, Any] | None = None
    days: list[dict[str, Any]] = []
    warnings: list[str] = []

    def ensure_node_payload(day_payload: dict[str, Any], rule: dict[str, Any]) -> dict[str, Any]:
        for node in day_payload["nodes"]:
            if node["node_type"] == rule["node_type"]:
                return node
        description_parts = []
        workflow_text = normalize_text(workflow_row[rule["column"] - 1]) if len(workflow_row) >= rule["column"] else ""
        time_text = normalize_text(time_row[rule["column"] - 1]) if len(time_row) >= rule["column"] else ""
        if workflow_text:
            description_parts.append(workflow_text)
        if time_text:
            description_parts.append(f"建议时间：{time_text}")
        node_payload = {
            "node_type": rule["node_type"],
            "title": rule["title"],
            "description": "\n".join(description_parts).strip(),
            "sort_order": rule["sort_order"],
            "msg_type": rule["msg_type"],
            "content": "",
            "variables_json": {},
            "status": "draft",
            "enabled": True,
        }
        day_payload["nodes"].append(node_payload)
        return node_payload

    for row in rows[header_idx + 3:]:
        first_col = normalize_text(row[0] if len(row) > 0 else "")
        second_col = normalize_text(row[1] if len(row) > 1 else "")
        third_col = normalize_text(row[2] if len(row) > 2 else "")

        if WEEK_PATTERN.match(first_col):
            week_label = first_col
            # 同一行可能同时有周标记和"第X天"，不要跳过
            if not second_col or not DAY_PATTERN.match(second_col):
                continue

        day_match = DAY_PATTERN.match(second_col)
        if day_match:
            current_day = {
                "day_number": chinese_to_int(day_match.group(1)),
                "week_label": week_label,
                "day_title": second_col,
                "day_focus": "",
                "nodes": [],
            }
            for rule in NODE_COLUMN_RULES:
                if len(row) < rule["column"]:
                    cell_text = ""
                else:
                    cell_text = normalize_text(row[rule["column"] - 1])
                node_payload = ensure_node_payload(current_day, rule)
                if cell_text:
                    node_payload["content"] = cell_text
                    node_payload["status"] = "ready"
                elif rule["node_type"] in DEFAULT_NODE_CONTENT:
                    node_payload["content"] = DEFAULT_NODE_CONTENT[rule["node_type"]]
                    node_payload["status"] = "ready"
            days.append(current_day)
            continue

        if not current_day:
            continue

        row_has_content = False
        for rule in NODE_COLUMN_RULES:
            if len(row) < rule["column"]:
                continue
            cell_text = normalize_text(row[rule["column"] - 1])
            if not cell_text:
                continue
            row_has_content = True
            node_payload = ensure_node_payload(current_day, rule)
            block_title = f"{third_col}\n" if third_col else ""
            block = f"{block_title}{cell_text}".strip()
            node_payload["content"] = f"{node_payload['content']}\n\n{block}".strip() if node_payload["content"] else block
            node_payload["status"] = "ready"

        if row_has_content and not third_col:
            warnings.append(f"第 {current_day['day_number']} 天存在未标注标签的补充内容，已按原文追加")

    if not days:
        raise HTTPException(400, "sheet1 中未识别到任何“第X天”数据行")

    title_cell = normalize_text(rows[0][0]) if rows and rows[0] else ""
    plan_name, stage, topic = parse_stage_and_topic(sheet_name, title_cell=title_cell)
    return {
        "plan_name": plan_name,
        "stage": stage,
        "topic": topic,
        "description": f"由 sheet1《{sheet_name}》导入生成",
        "day_count": len(days),
        "days": days,
        "warnings": warnings,
    }


async def parse_operation_plan_sheet1(file: UploadFile) -> dict[str, Any]:
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(400, "仅支持导入 .xlsx 文件")

    content = await file.read()
    if not content:
        raise HTTPException(400, "上传文件为空")

    try:
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"无法解析 Excel 文件：{exc}") from exc

    if not workbook.sheetnames:
        raise HTTPException(400, "Excel 文件中没有可用工作表")

    sheet = workbook[workbook.sheetnames[0]]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    return parse_sheet1_rows(rows, sheet.title)


JSON_STAGE_RULE_MAP = {
    "morning_greeting_points": {
        "node_type": "morning_breakfast",
        "title": "早餐提醒+早安+积分发布",
        "field": "morning_text",
    },
    "pre_lunch_knowledge": {
        "node_type": "before_lunch_content",
        "title": "午餐前：知识点预告-讲科学+给方法",
        "field": "pre_lunch_knowledge_text",
    },
    "lunch_reminder": {
        "node_type": "lunch_reminder",
        "title": "午餐提醒",
        "field": "lunch_reminder_text",
    },
    "afternoon_meal_review": {
        "node_type": "afternoon_review",
        "title": "下午餐评+互动+答疑",
        "field": "afternoon_meal_review_text",
    },
    "dinner_reminder": {
        "node_type": "dinner_reminder",
        "title": "晚餐提醒+答疑",
        "field": "dinner_reminder_text",
    },
    "dinner_meal_review": {
        "node_type": "evening_review",
        "title": "晚餐餐评+互动+答疑",
        "field": "dinner_meal_review_text",
    },
    "night_summary_preview": {
        "node_type": "night_summary",
        "title": "晚安+总结+明天预告",
        "field": "night_summary_text",
    },
}


def split_stage_topic(raw: str) -> tuple[str, str, str]:
    text = normalize_text(raw)
    if "：" in text:
        stage, topic = text.split("：", 1)
    elif ":" in text:
        stage, topic = text.split(":", 1)
    else:
        stage, topic = text, text
    return text, stage.strip(), topic.strip()


def build_json_description(payload: dict[str, Any]) -> str:
    blocks: list[str] = []
    source_file = normalize_text(payload.get("source_file"))
    if source_file:
        blocks.append(f"导入源文件：{source_file}")

    notes = [normalize_text(item) for item in payload.get("important_notes", []) if normalize_text(item)]
    if notes:
        blocks.append("重要说明：\n" + "\n".join(f"- {item}" for item in notes))

    operator_rules = [normalize_text(item) for item in payload.get("operator_fixed_rules", []) if normalize_text(item)]
    if operator_rules:
        blocks.append("医护执行固定规则：\n" + "\n".join(f"- {item}" for item in operator_rules))

    weekly_events = payload.get("weekly_events", [])
    if isinstance(weekly_events, list) and weekly_events:
        event_lines = []
        for event in weekly_events:
            if not isinstance(event, dict):
                continue
            day_rule = normalize_text(event.get("day_rule"))
            action = normalize_text(event.get("action"))
            time = normalize_text(event.get("time"))
            line = " / ".join(part for part in [day_rule, time, action] if part and part != "未写明确切时间")
            if line:
                event_lines.append(f"- {line}")
        if event_lines:
            blocks.append("固定周事件：\n" + "\n".join(event_lines))

    return "\n\n".join(blocks).strip()


def build_json_day_focus(day_payload: dict[str, Any]) -> str:
    lines: list[str] = []
    week_label = normalize_text(day_payload.get("week_label"))
    phase_name = normalize_text(day_payload.get("phase_name"))
    knowledge_topic = normalize_text(day_payload.get("knowledge_topic"))
    required_actions = [
        normalize_text(item) for item in day_payload.get("required_actions", []) if normalize_text(item)
    ]

    if week_label:
        lines.append(f"周次：{week_label}")
    if phase_name:
        lines.append(f"阶段任务：{phase_name}")
    if knowledge_topic:
        lines.append(f"今日知识主题：{knowledge_topic}")
    if required_actions:
        lines.append("必做动作：")
        lines.extend(f"- {item}" for item in required_actions)
    return "\n".join(lines).strip()


def build_json_node_description(template_payload: dict[str, Any]) -> str:
    parts: list[str] = []
    scheduled_time = normalize_text(template_payload.get("scheduled_time"))
    owner = normalize_text(template_payload.get("owner"))
    fixed_action = normalize_text(template_payload.get("fixed_action"))
    source_basis = normalize_text(template_payload.get("source_basis"))
    if scheduled_time:
        parts.append(f"建议时间：{scheduled_time}")
    if owner:
        parts.append(f"执行角色：{owner}")
    if fixed_action:
        parts.append(fixed_action)
    if source_basis:
        parts.append(f"来源依据：{source_basis}")
    return "\n".join(parts).strip()


def parse_operation_plan_json_bytes(content: bytes) -> dict[str, Any]:
    try:
        payload = json.loads(content.decode("utf-8"))
    except UnicodeDecodeError:
        payload = json.loads(content.decode("utf-8-sig"))
    except json.JSONDecodeError as exc:
        raise HTTPException(400, f"无法解析 JSON 文件：{exc}") from exc

    if not isinstance(payload, dict):
        raise HTTPException(400, "JSON 导入文件必须是对象结构")

    # 兼容两种格式：
    # 1) 通用导出格式：plan_name + days[].nodes[]
    # 2) 原有 v2 格式：extraction_scope + days[].morning_text 等
    plan_name_raw = normalize_text(payload.get("plan_name"))
    extraction_scope = normalize_text(payload.get("extraction_scope"))
    if not plan_name_raw and not extraction_scope:
        raise HTTPException(400, "JSON 导入文件缺少 plan_name 或 extraction_scope")

    if plan_name_raw:
        plan_name = plan_name_raw
        stage = normalize_text(payload.get("stage")) or plan_name_raw
        topic = normalize_text(payload.get("topic")) or plan_name_raw
    else:
        plan_name, stage, topic = split_stage_topic(extraction_scope)

    days_raw = payload.get("days", [])
    if not isinstance(days_raw, list) or not days_raw:
        raise HTTPException(400, "JSON 导入文件缺少 days")

    warnings: list[str] = []

    # 检测格式：如果 days[0] 有 nodes 数组，走通用格式解析
    first_day = days_raw[0] if days_raw else {}
    if isinstance(first_day, dict) and isinstance(first_day.get("nodes"), list):
        days = _parse_generic_json_days(days_raw, warnings)
    else:
        days = _parse_v2_json_days(days_raw, payload, warnings)

    # 优先用 plan_name 作为 description（通用格式没有 build_json_description）
    description = normalize_text(payload.get("description"))
    if not description and extraction_scope:
        description = build_json_description(payload)

    return {
        "plan_name": plan_name,
        "stage": stage,
        "topic": topic,
        "description": description,
        "day_count": len(days),
        "days": sorted(days, key=lambda item: item["day_number"]),
        "warnings": warnings,
    }


def _parse_generic_json_days(
    days_raw: list[dict], warnings: list[str]
) -> list[dict[str, Any]]:
    """解析通用导出格式：days[].nodes[] 数组"""
    days: list[dict[str, Any]] = []
    for raw_day in days_raw:
        if not isinstance(raw_day, dict):
            continue
        day_number = raw_day.get("day")
        if not isinstance(day_number, int):
            continue

        nodes_raw = raw_day.get("nodes", [])
        nodes: list[dict[str, Any]] = []
        for raw_node in nodes_raw:
            if not isinstance(raw_node, dict):
                continue
            content = raw_node.get("content", "")
            nodes.append({
                "node_type": normalize_text(raw_node.get("node_type")) or "custom",
                "title": normalize_text(raw_node.get("title")) or "",
                "description": normalize_text(raw_node.get("description")) or "",
                "sort_order": int(raw_node.get("sort_order") or len(nodes) + 1) * 10,
                "msg_type": normalize_text(raw_node.get("msg_type")) or "markdown",
                "content": normalize_text(content),
                "variables_json": raw_node.get("variables_json") if isinstance(raw_node.get("variables_json"), dict) else {},
                "status": "ready" if content else "draft",
                "enabled": True,
            })

        if not nodes:
            warnings.append(f"第 {day_number} 天没有任何节点")

        days.append({
            "day_number": day_number,
            "week_label": normalize_text(raw_day.get("week_label")),
            "day_title": normalize_text(raw_day.get("day_title")) or f"第{day_number}天",
            "day_focus": normalize_text(raw_day.get("day_focus")),
            "nodes": nodes,
        })
    return days


def _parse_v2_json_days(
    days_raw: list[dict], payload: dict, warnings: list[str]
) -> list[dict[str, Any]]:
    """解析原有 v2 格式：days[].morning_text 等扁平字段"""
    stage_templates = payload.get("stage_templates", [])

    template_by_key: dict[str, dict[str, Any]] = {}
    for item in stage_templates:
        if not isinstance(item, dict):
            continue
        key = normalize_text(item.get("stage_key"))
        if key:
            template_by_key[key] = item

    days: list[dict[str, Any]] = []

    for raw_day in days_raw:
        if not isinstance(raw_day, dict):
            continue
        day_number = raw_day.get("day")
        if not isinstance(day_number, int):
            raise HTTPException(400, "JSON days[].day 必须是整数")

        day_payload = {
            "day_number": day_number,
            "week_label": normalize_text(raw_day.get("week_label")),
            "day_title": f"第{day_number}天",
            "day_focus": build_json_day_focus(raw_day),
            "nodes": [],
        }

        for template_key, mapping in JSON_STAGE_RULE_MAP.items():
            template_payload = template_by_key.get(template_key, {})
            field_name = mapping["field"]
            content_blocks = []
            main_text = normalize_text(raw_day.get(field_name))
            if main_text:
                content_blocks.append(main_text)

            if template_key == "pre_lunch_knowledge":
                extra_text = normalize_text(raw_day.get("knowledge_extra_text"))
                evidence_text = normalize_text(raw_day.get("knowledge_evidence_text"))
                if extra_text:
                    content_blocks.append(extra_text)
                if evidence_text:
                    content_blocks.append(evidence_text)

            if not content_blocks:
                warnings.append(f"第 {day_number} 天缺少 {mapping['title']} 内容，已创建空节点")

            day_payload["nodes"].append(
                {
                    "node_type": mapping["node_type"],
                    "title": normalize_text(template_payload.get("stage_name")) or mapping["title"],
                    "description": build_json_node_description(template_payload),
                    "sort_order": int(template_payload.get("stage_order") or len(day_payload["nodes"]) + 1) * 10,
                    "msg_type": "markdown",
                    "content": "\n\n".join(block for block in content_blocks if block).strip(),
                    "variables_json": {
                        "phase_name": normalize_text(raw_day.get("phase_name")),
                        "knowledge_topic": normalize_text(raw_day.get("knowledge_topic")),
                    },
                    "status": "ready" if content_blocks else "draft",
                    "enabled": True,
                }
            )

        days.append(day_payload)
    return days


async def parse_operation_plan_file(
    file: UploadFile, import_type: str = 'daily_sop',
) -> dict[str, Any]:
    if import_type == 'points_campaign':
        from .operation_plan_import_points import parse_points_campaign_excel
        content = await file.read()
        if not content:
            raise HTTPException(400, "上传文件为空")
        try:
            workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
        except Exception as exc:
            raise HTTPException(400, f"无法解析 Excel 文件：{exc}") from exc
        return parse_points_campaign_excel(workbook)

    filename = (file.filename or "").lower()
    if filename.endswith(".json"):
        content = await file.read()
        if not content:
            raise HTTPException(400, "上传文件为空")
        return parse_operation_plan_json_bytes(content)
    return await parse_operation_plan_sheet1(file)
